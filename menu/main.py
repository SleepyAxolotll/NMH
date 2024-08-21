from fastapi import FastAPI, UploadFile
from fastapi.responses import JSONResponse, FileResponse
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os

app = FastAPI()

# Specify the directory to save the Excel files
SAVE_DIRECTORY = "excel_files"

@app.post("/process_menu")
async def process_menu(file_upload: UploadFile):
    try:
        if not os.path.exists(SAVE_DIRECTORY):
            os.makedirs(SAVE_DIRECTORY)
        excel_file_path = os.path.join(SAVE_DIRECTORY, "menu.xlsx")

        wb = Workbook()
        ws = wb.active
        html_content = await file_upload.read()
        soup = BeautifulSoup(html_content, 'html.parser')

        date_wrappers = soup.find_all(class_="day-name")
        meal_names = [meal_name.h3.get_text(strip=True) for meal_name in soup.find_all(class_="meal-name")]
        ul_elements = soup.find_all('ul')

        food_times_list = [x.getText(strip=True) for x in date_wrappers]
        for index, value in enumerate(food_times_list, start=2):
            ws.cell(row=index, column=1, value=value)

        meals_list = ["LUNCH", "DINNER"]
        for col, meal in enumerate(meals_list, start=2):
            ws.cell(row=1, column=col, value=meal)

        day_tracker = 2
        meal_name_counter = 0
        unwanted_food = ["vegetable", "allergens", "traceallergens", "rice", "pasta"]
    
        for ul in ul_elements:
            current_meal_name = meal_names[meal_name_counter]
            meal_name_counter += 1

            if "BREAKFAST" not in current_meal_name:
                filtered_food_elements = [
                    food_item.find('div').get_text(strip=True)
                    for food_item in ul.find_all('li', class_='food')
                    if not any(unwanted_word in food_item.get_text(strip=True).lower() for unwanted_word in unwanted_food)
                ]

                list_string = ', '.join(map(str, filtered_food_elements))

            if "LUNCH" in current_meal_name:
                ws.cell(row=day_tracker, column=2, value=list_string)
            elif "DINNER" in current_meal_name:
                ws.cell(row=day_tracker, column=3, value=list_string)
                day_tracker += 1

        wb.save(excel_file_path)
        return FileResponse(excel_file_path, filename="menu.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        return JSONResponse(status_code=500, content={"message": f"An error occurred: {str(e)}"})
